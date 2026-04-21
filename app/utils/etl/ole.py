import io
from app.utils.etl.base import BaseParser


class OLEExcelParser(BaseParser):
    def parse(self, data: bytes) -> list[dict]:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            ws = wb.active
            headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(max_row=1))]
            records = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                if not self.validate_record(row_dict):
                    continue
                records.append({
                    "nombre": str(row_dict.get("PROGRAMA", "") or "").strip(),
                    "salarioEntrada": int(row_dict.get("SALARIO_INGRESO", 0) or 0),
                    "salarioMedio": int(row_dict.get("SALARIO_MEDIO", 0) or 0),
                    "tasaEmpleabilidad12m": float(row_dict.get("EMPLEABILIDAD_12M", 0) or 0),
                    "fuenteSalario": "OLE",
                })
            return records
        except Exception as exc:
            raise ValueError(f"OLE Excel parse error: {exc}") from exc

    def validate_record(self, record: dict) -> bool:
        return bool(record.get("PROGRAMA")) and bool(record.get("SALARIO_INGRESO"))
